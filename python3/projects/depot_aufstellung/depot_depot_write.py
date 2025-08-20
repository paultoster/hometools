#
# datensatz anzeigen
#
#
#
import os
import sys

import openpyxl
from pathlib import Path

tools_path = os.getcwd() + "\\.."
if (tools_path not in sys.path):
    sys.path.append(tools_path)
# endif

# Hilfsfunktionen
import hfkt_def as hdef
import hfkt_date_time as hdate
# import hfkt_str as hstr

import depot_gui


def ods_ausgabe_mit_depot_wahl(rd):
    '''

    :param rd:
    :return: status =  ods_ausgabe_mit_depot_wahl(rd)
    '''
    
    status = hdef.OKAY
    
    # Kontoauswählen
    runflag = True
    while (runflag):
        
        (index, auswahl) = depot_gui.auswahl_depot(rd)
        
        if index < 0:
            return status
        elif auswahl in rd.ini.ddict[rd.par.INI_DEPOT_DATA_DICT_NAMES_NAME]:
            
            rd.log.write(f"depot  \"{auswahl}\" ausgewählt")
            break
        else:
            status = hdef.NOT_OKAY
            errtext = f"Depot Auswahl: {auswahl} nicht bekannt"
            rd.log.write_err(errtext, screen=rd.par.LOG_SCREEN_OUT)
            return status
        # endif
    # endwhile
    
    # Konto data in ini
    depot_dict = rd.data[auswahl].ddict
    depot_obj = rd.data[auswahl].obj

    
    file_name = hdate.get_name_by_dat_time(depot_obj.depot_name+"_","")+".xlsx"
    # start ods-Output
    workbook = openpyxl.Workbook()
    
    worksheet = workbook.active
    
    # Übersichtsdaten von allen  WPs
    worksheet.title = depot_obj.depot_name
    
    (data_lliste, header_liste, _ , _) = depot_obj.get_depot_daten_sets_overview(0)
    
    if depot_obj.status != hdef.OKAY:
        status = depot_obj.status
        rd.log.write_err(depot_obj.errtext, screen=rd.par.LOG_SCREEN_OUT)
        depot_obj.reset_status()
        workbook.save(file_name)
        return status
    # end if
    
    row_num = 1
    for i,header in enumerate(header_liste):
        col_num = i+1
        worksheet.cell(row=row_num,column=col_num,value=header)
    # end if
    for data_liste in data_lliste:
        row_num += 1
        for i,data in enumerate(data_liste):
            col_num = i + 1
            worksheet.cell(row=row_num, column=col_num, value=data)
        # end for
    # end for
    
    # isin_liste
    isin_liste = depot_obj.get_isin_liste()
    
    for isin in isin_liste:
        
        worksheet = workbook.create_sheet(title=isin)
        workbook.active = workbook[isin]
        
        (data_lliste, header_liste, _ , _) = depot_obj.get_depot_daten_sets_isin(isin)
        
        if depot_obj.status != hdef.OKAY:
            status = depot_obj.status
            rd.log.write_err(depot_obj.errtext, screen=rd.par.LOG_SCREEN_OUT)
            depot_obj.reset_status()
            workbook.save(file_name)
            return status
        # end if
        
        row_num = 1
        for i,header in enumerate(header_liste):
            col_num = i+1
            worksheet.cell(row=row_num,column=col_num,value=header)
        # end if
        for data_liste in data_lliste:
            row_num += 1
            for i,data in enumerate(data_liste):
                col_num = i + 1
                worksheet.cell(row=row_num, column=col_num, value=data)
            # end for
        # end for
    # end for
    
    workbook.save(file_name)
    
    
    
    rd.log.write(f"excel-File-Name: {os.getcwd()}\\{file_name}",screen=rd.par.LOG_SCREEN_OUT)
    
    return status
# enddef
