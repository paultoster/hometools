#
# datensatz anzeigen und Kategorien bearbeiten
#
#
#
import os
import sys

import openpyxl

tools_path = os.getcwd() + "\\.."
if (tools_path not in sys.path):
    sys.path.append(tools_path)
# endif

# Hilfsfunktionen
import tools.hfkt_def as hdef
import tools.hfkt_date_time as hdate
# import tools.hfkt_str as hstr
import tools.hfkt_tvar as htvar
import tools.hfkt_type as htype
import tools.hfkt_file_path as hfile

import depot_kategorie_auswertung_class

def excel_auswertung(rd):
    
    jahr = rd.ini.dict_tvar[rd.par.INI_KONTO_AUSWERTUNG_JAHR_NAME].val
    auswert_path = rd.ini.dict_tvar[rd.par.INI_AUSWERT_PATH].val

    (tval_min_time,tval_max_time) = get_limits_epocht_time_of_year(jahr)
    
    if len(rd.ini.ddict[rd.par.INI_KONTO_DATA_LIST_NAMES_NAME]) == 0:
        rd.log.write_err("anzeige  kategorie: kein Konto eingerichtet", screen=rd.par.LOG_SCREEN_OUT)
        return hdef.NOT_OKAY
    # end if
    
    kont_name = rd.ini.ddict[rd.par.INI_KONTO_DATA_LIST_NAMES_NAME][0]
    konto_obj = rd.konto_dict[kont_name].konto_obj
    
    header_list = [konto_obj.par.KONTO_DATA_NAME_BUCHDATUM,
                   konto_obj.par.KONTO_DATA_NAME_WER,
                   konto_obj.par.KONTO_DATA_NAME_WERT,
                   konto_obj.par.KONTO_DATA_NAME_COMMENT,
                   konto_obj.par.KONTO_DATA_NAME_KATEGORIE]
    type_list = ["dat", "str", "cent", "str", "str"]
    
    # Erstelle Auswert Klassenobjekt
    # zur headerlist kommt noch der Kontoname:
    katauswertfunc = depot_kategorie_auswertung_class.KategorieAuswertungClass(konto_obj.par,
                                                                         [konto_obj.par.KONTO_NAME] + header_list,
                                                                         ["str"] + type_list,
                                                                         rd.allg.katfunc,
                                                                         jahr)
    
    # Loop over each konto
    for kont_name in rd.ini.ddict[rd.par.INI_KONTO_DATA_LIST_NAMES_NAME]:
        
        konto_obj = rd.konto_dict[kont_name].konto_obj
        
        # print(f"vor: {header_list =}")
        
        ttable = konto_obj.get_timedepend_data_set_dict_ttable(header_list, type_list,
                                                               tval_min_time, tval_max_time,
                                                               with_start=False)
        
        # print(f"nach: {header_list =}")
        
        if konto_obj.status != hdef.OKAY:
            rd.log.write_err("anzeige  kategorie " + konto_obj.errtext, screen=rd.par.LOG_SCREEN_OUT)
            konto_obj.reset_status()
            return konto_obj.status
        # end if
        
        add_row_liste = [kont_name for i in range(ttable.ntable)]
        ttable = htvar.add_row_liste_to_table(ttable, konto_obj.par.KONTO_NAME, add_row_liste, "str")
        # print(f"nach1: {header_list =}")
        
        status = katauswertfunc.add_ttable(ttable)
        # print(f"nach2: {header_list =}")
        if status != hdef.OKAY:
            rd.log.write_err("anzeige  kategorie add ttable \n" + katauswertfunc.errtext, screen=rd.par.LOG_SCREEN_OUT)
            katauswertfunc.reset_status()
            return status
        # end if
        # print(f"nach3: {header_list =}")
    
    # end for
    
    # run Auswertung
    #===============
    katauswertfunc.run_auswert()
    
    (ttable_ueberblick,index_ttable_ueberblick) = katauswertfunc.get_auswert_ueberblick()
    if katauswertfunc.status != hdef.OKAY:
        status = katauswertfunc.status
        rd.log.write_err("anzeige  kategorie add ttable \n" + katauswertfunc.errtext, screen=rd.par.LOG_SCREEN_OUT)
        katauswertfunc.reset_status()
        return status
    # end if

    (kat_liste,ttable_einzel_liste,ttable_color_index_liste) = katauswertfunc.get_auswert_einzel_liste()
    if katauswertfunc.status != hdef.OKAY:
        status = katauswertfunc.status
        rd.log.write_err("anzeige  kategorie add ttable \n" + katauswertfunc.errtext, screen=rd.par.LOG_SCREEN_OUT)
        katauswertfunc.reset_status()
        return status
    # end if
    
    
    
    if not os.path.isdir(auswert_path):
        try:
            os.mkdir(auswert_path)
        except OSError as e:
            raise Exception(f"Error: {e}")
        # end try
    # end if
    file_body         = hdate.get_name_by_dat_time("Kontoauswertung_" + str(jahr) + "_", "")
    file_name         = hfile.set_pfe(auswert_path, file_body, "xlsx")
    
    status = bilde_ods_File(file_name,ttable_ueberblick,index_ttable_ueberblick,kat_liste,ttable_einzel_liste,ttable_color_index_liste)
    
    return status
# end def
def bilde_ods_File(file_name, ttable_ueberblick,index_ttable_ueberblick,kat_liste,ttable_einzel_liste,ttable_color_index_liste):
    '''
    
    :param file_name:
    :param ttable_ueberblick:
    :return: status = bilde_ods_File(file_name, ttable_ueberblick)
    '''
    front_page_title = "Zusammenfassung"
    max_col_wdth     = 120
    # start ods-Output
    workbook = openpyxl.Workbook()
    
    worksheet = workbook.active
    
    # Übersichtsdaten von allen  WPs
    worksheet.title = front_page_title
    
    ft = openpyxl.styles.fonts.Font(bold=True)
    al = openpyxl.styles.Alignment(horizontal='right',wrap_text=True)
    
    # border_thick = openpyxl.styles.Side(style='thick')
    # border_thin = openpyxl.styles.Side(style='thin')
    borded_dashed = openpyxl.styles.Side(style='dashed')
    # border_dotted = openpyxl.styles.Side(style='dotted')
    
    bd = openpyxl.styles.Border(top=borded_dashed, left=borded_dashed, right=borded_dashed, bottom=borded_dashed)
    row_num = 0
    for i,header in enumerate(ttable_ueberblick.names):
        worksheet[htype.excel_calc_alph_num(row_num,i)] = header
        worksheet[htype.excel_calc_alph_num(row_num, i)].font = ft
        worksheet[htype.excel_calc_alph_num(row_num, i)].alignment  = al
        worksheet[htype.excel_calc_alph_num(row_num, i)].border = bd
    # end if
    
    for irow,data_liste in enumerate(ttable_ueberblick.table):
        row_num += 1
        for icol,data in enumerate(data_liste):
            worksheet[htype.excel_calc_alph_num(row_num, icol)] = data
            color = htype.get_excel_color_by_index(index_ttable_ueberblick.table[irow][icol])
            fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
            worksheet[htype.excel_calc_alph_num(row_num, icol)].fill = fill
            worksheet[htype.excel_calc_alph_num(row_num, icol)].alignment = al
            worksheet[htype.excel_calc_alph_num(row_num, icol)].border = bd
            # end for
    # end for
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        max_length = min(max_length,max_col_wdth)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        worksheet.column_dimensions[col_letter].width = max_length + 5
    # end for
    
    # fill_color = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # cell.fill = fill_color
    
    # Einzelwerte:
    #-------------
    for index,kat in enumerate(kat_liste):
        
        worksheet = workbook.create_sheet(title=kat)
        workbook.active = workbook[kat]
        
        ttable = ttable_einzel_liste[index]
        ttable_color_index = ttable_color_index_liste[index]
        
        # Header
        row_num = 0
        for icol,header in enumerate(ttable.names):
            worksheet[htype.excel_calc_alph_num(row_num,icol)] = header
            worksheet[htype.excel_calc_alph_num(row_num, icol)].font = ft
            worksheet[htype.excel_calc_alph_num(row_num, icol)].alignment  = al
            worksheet[htype.excel_calc_alph_num(row_num, icol)].border = bd
        # end if

        for irow,data_liste in enumerate(ttable.table):
            row_num += 1
            for icol,data in enumerate(data_liste):
                worksheet[htype.excel_calc_alph_num(row_num, icol)] = data
                color = htype.get_excel_color_by_index(ttable_color_index.table[irow][icol])
                fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
                worksheet[htype.excel_calc_alph_num(row_num, icol)].fill = fill
                worksheet[htype.excel_calc_alph_num(row_num, icol)].alignment = al
                worksheet[htype.excel_calc_alph_num(row_num, icol)].border = bd
                # end for
        # end for
        
        # Adjust column widths
        for col in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            max_length = min(max_length, max_col_wdth)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max_length + 5
            # print(f"{max_length = }")
        # end for
    # end for
    workbook.active = workbook[front_page_title]
    workbook.save(file_name)
    
    os.startfile(file_name)
    
    status = hdef.OKAY
    
    return status
# end def

def get_limits_epocht_time_of_year(str_year):
    '''
    
    :param str_jear:
    :return: (min_epoch_time, max_epoch_time) = get_limits_epocht_time_of_year(str_year)
    '''
    
    str_year_next = str(int(str_year) + 1)
    
    (okay,wert) = htype.type_transform_yearStr(str_year,"dat")
    
    if okay != hdef.OKAY:
        raise Exception(f"get_limits_epocht_time_of_year: {str_year = } kann nicht in epochaler Zeit gewandelt werten")
    # end if
    tval_min_epoch_time = htvar.build_val("min_epoch_time",wert,"dat")
    
    (okay,wert) = htype.type_transform_yearStr(str_year_next,"dat")
    if okay != hdef.OKAY:
        raise Exception(f"get_limits_epocht_time_of_year: {str_year_next = } kann nicht in epochaler Zeit gewandelt werten")
    # end if
    tval_max_epoch_time = htvar.build_val("max_epoch_time", wert, "dat")
    
    return (tval_min_epoch_time, tval_max_epoch_time)
# end def
