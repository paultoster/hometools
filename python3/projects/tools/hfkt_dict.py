
"""
keylist = find_keys_of_dict_value_as_list(ddict,value)

key = find_first_key_dict_value(ddict,value) if not in value = None

(status,errtext,ddict) = proof_transform_ddict(ddict, proof_liste)
    
    Beispiel:
    ddict["wert"]  = "10,34"
    ddict["name"]  = "Berthold"
    ddict["datum"] = "10.03.2004"
    
    1. proof_liste = ["wert","name","datum"]     hier wird geprüft, ob vorhanden
    2. proof_liste = [("wert","euroStrK"),("name","str"),("datum","datStrP")]
                                                 hier wird der Type geprüft, ob der type stimmt (siehe hfkt_type.py type_proof())
    3. proof_liste = [("wert","euroStrK","euro"),
                      ("name","str","str"),
                      ("datum","datStrP","dat"2")]
                                                 hier wird der Type geprüft und transformiert "euroStrK" => "euro"
     

(status,errtext,dict_tvar) = proof_transform_ddict_to_tvar(ddict, proof_liste)


    Beispiel:
    ddict["wert"]  = "10,34"
    ddict["name"]  = "Berthold"
    ddict["datum"] = "10.03.2004"

    1. proof_liste = ["wert","name","datum"]     hier wird geprüft, ob vorhanden
    2. proof_liste = [("wert","euroStrK"),("name","str"),("datum","datStrP")]
                                                 hier wird der Type geprüft, ob der type stimmt (siehe hfkt_type.py type_proof())
    3. proof_liste = [("wert","euroStrK","euro"),
                      ("name","str","str"),
                      ("datum","datStrP","dat")]
                                                 hier wird der Type geprüft und transformiert "euroStrK" => "euro"

    output ist
     dict_tvar["wert"] = TVal("wert",10.34,"euro")
     dict_tvar["name"] = TVal("name","Berthold","str")
     dict_tvar["datum"] = TVal("datum",2787388348,"dat")
    

"""
import os
import openpyxl

if os.path.isfile('hfkt_def.py'):
    import hfkt_def as hdef
    import hfkt_type as htype
    import hfkt_dict as hdict
    import hfkt_tvar as htvar
    import hfkt_file_path as hfp
else:
    import tools.hfkt_def as hdef
    import tools.hfkt_type as htype
    import tools.hfkt_dict as hdict
    import tools.hfkt_tvar as htvar
    import tools.hfkt_file_path as hfp
# end if



def find_keys_of_dict_value_as_list(ddict,value):
    keys = [key for key, val in ddict.items() if val == value]
    return keys
# end def
def find_first_key_dict_value(ddict,value):
    keys = find_keys_of_dict_value_as_list(ddict,value)
    if len(keys):
        return keys[0]
    else:
        return None
# end def
def proof_transform_ddict(ddict, proof_liste):
    '''
    
    Beispiel:
    ddict["wert"]  = "10,34"
    ddict["name"]  = "Berthold"
    ddict["datum"] = "10.03.2004"
    
    1. proof_liste = ["wert","name","datum"]     hier wird geprüft, ob vorhanden
    2. proof_liste = [("wert","euroStrK"),("name","str"),("datum","datStrP")]
                                                 hier wird der Type geprüft, ob der type stimmt (siehe hfkt_type.py type_proof())
    3. proof_liste = [("wert","euroStrK","euro"),
                      ("name","str","str"),
                      ("datum","datStrP","dat"2")]
                                                 hier wird der Type geprüft und transformiert "euroStrK" => "euro"
    4. proof_liste = [("wert","euroStrK","euro","0,00")]  gibt einen default-Wert an
     
    
    :param ddict:
    :param proof_liste:
    :return: (status,errtext,ddict) = proof_transform_ddict(ddict, proof_liste)
    '''
    status = hdef.OKAY
    errtext = ""
    for index,item in enumerate(proof_liste):
        
        proof_name = False
        proof_type = False
        trans_type = False
        default_val = False
        if isinstance(item,str):
            proof_name = True
            name       = item
        elif isinstance(item,list) or isinstance(item,tuple):
            n = len(item)
            if n == 0:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: proof_liste wird nicht erkannt proof_liste[{index}] = {item}"
                return (status, errtext,ddict)
            # end if
            if (n > 0) and isinstance(item[0],str):
                proof_name = True
                name = item[0]
            # end if
            if (n > 1) and isinstance(item[1],str):
                proof_type = True
            # end if
            if (n > 2) and isinstance(item[2],str):
                trans_type = True
            # end if
            if n > 3:
                default_val = True
            # end if
        else:
            status = hdef.NOT_OKAY
            errtext = f"proof_transform_ddict: proof_liste wird nicht erkannt proof_liste[{index}] = {item}"
            return (status,errtext,ddict)
        # end if
        
        if proof_name:
            if default_val:
                if name in ddict:
                    defualt_val = False
                else:
                    ddict[name] = item[3]
                # end if
            elif  (name not in ddict):
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: proof_liste[{index}] = {name} ist nicht in dictionary"
                return (status, errtext,ddict)
            # end if
        # end if
        if trans_type:
            [okay, wert] = htype.type_transform(ddict[name], item[1], item[2])
            if okay == hdef.OKAY:
                ddict[name]  = wert
            else:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: ddict[{name}] = {ddict[name]} kann in dict nicht transformiert werden type: {item[1]} => {item[2]}"
                return (status, errtext, ddict)
            # end if
        else:
            [okay, wert] = htype.type_proof(ddict[name], item[1])
            if okay != hdef.OKAY:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: ddict[{name}] = {ddict[name]}  type: {item[1]} stimmt nicht"
                return (status, errtext, ddict)
            # end if
        # end if
    # end for
    return (status, errtext, ddict)
# end def
def proof_transform_ddict_to_tvar(ddict, proof_liste):
    '''

    Beispiel:
    ddict["wert"]  = "10,34"
    ddict["name"]  = "Berthold"
    ddict["datum"] = "10.03.2004"

    1. proof_liste = ["wert","name","datum"]     hier wird geprüft, ob vorhanden
    2. proof_liste = [("wert","euroStrK"),("name","str"),("datum","datStrP")]
                                                 hier wird der Type geprüft, ob der type stimmt (siehe hfkt_type.py type_proof())
    3. proof_liste = [("wert","euroStrK","euro"),
                      ("name","str","str"),
                      ("datum","datStrP","dat")]
                                                 hier wird der Type geprüft und transformiert "euroStrK" => "euro"
    4. proof_liste = [("wert","euroStrK","euro","0,0"),  vierte Position ist defaultwert
                      ("name","str","str"),
                      ("datum","datStrP","dat")]
                                                 hier wird der Type geprüft und transformiert "euroStrK" => "euro"

    output ist
     dict_tvar["wert"] = TVal("wert",10.34,"euro")
     dict_tvar["name"] = TVal("name","Berthold","str")
     dict_tvar["datum"] = TVal("datum",2787388348,"dat")
    
    :param ddict:
    :param proof_liste:
    :return: (status,errtext,dict_tvar) = proof_transform_ddict_to_TVar(ddict, proof_liste)
    '''
    status = hdef.OKAY
    errtext = ""
    dict_tvar = {}
    for index, item in enumerate(proof_liste):
        
        proof_name = False
        proof_type = False
        trans_type = False
        default_value = None
        if isinstance(item, str):
            proof_name = True
            name = item
        elif isinstance(item, list) or isinstance(item, tuple):
            n = len(item)
            if n == 0:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: proof_liste wird nicht erkannt proof_liste[{index}] = {item}"
                return (status, errtext, dict_tvar)
            # end if
            if (n > 0) and isinstance(item[0], str):
                proof_name = True
                name = item[0]
            # end if
            if (n > 1) and isinstance(item[1], str):
                proof_type = True
            # end if
            if (n > 2) and isinstance(item[2], str):
                trans_type = True
            # end if
            if (n > 3):
                default_value = item[3]
            # end if
        else:
            status = hdef.NOT_OKAY
            errtext = f"proof_transform_ddict: proof_liste wird nicht erkannt proof_liste[{index}] = {item}"
            return (status, errtext, dict_tvar)
        # end if
        
        if proof_name and (name not in ddict):
            if default_value is not None:
                ddict[name] = default_value
            else:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: proof_liste[{index}] = {name} ist nicht in dictionary"
                return (status, errtext, dict_tvar)
        # end if
        if trans_type:
            [okay, wert] = htype.type_transform(ddict[name], item[1], item[2])
            if okay == hdef.OKAY:
                dict_tvar[name] = htvar.build_val(name,wert,item[2])
            else:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: ddict[{name}] = {ddict[name]} kann in dict nicht transformiert werden type: {item[1]} => {item[2]}"
                return (status, errtext, dict_tvar)
            # end if
        else:
            [okay, wert] = htype.type_proof(ddict[name], item[1])
            if okay != hdef.OKAY:
                status = hdef.NOT_OKAY
                errtext = f"proof_transform_ddict: ddict[{name}] = {ddict[name]}  type: {item[1]} stimmt nicht"
                return (status, errtext, ddict)
            # end if
            dict_tvar[name] = htvar.build_val(name, wert, item[1])
        # end if
    # end for
    return (status, errtext, dict_tvar)
    # end def
def add_dict_to_dict(dict_base,dict_add):
    '''
    
    :param dict_base:
    :param dict_add:
    :return: dict_base = hfkt_dict.add_dict_to_dict(dict_base,dict_add)
    '''

    for key in dict_add.keys():
        dict_base[key] = dict_add[key]
    # end for
    
    return dict_base
# end def
def write_dict_list_in_ods_table(dict_list,titlename,filename):
    """
    :param dict_list: list aof dictionaries with key and value as string, interger or float
    :param filename:  Filename
    :return: (status,errtext) = write_dict_list_in_ods_table(dict_list,filename)
    """

    status = hdef.OKAY
    errtext = ""

    # start ods-Output
    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = titlename

    row_num = 1
    for i,header in enumerate(dict_list[0].keys()):
        col_num = i+1
        worksheet.cell(row=row_num,column=col_num,value=header)
    # end if
    for ddict in dict_list:
        row_num += 1
        for i,data in enumerate(ddict.values()):
            col_num = i + 1
            worksheet.cell(row=row_num, column=col_num, value=data)
        # end for
    # end for
    file_name = hfp.reset_ext(filename, "excel")
    workbook.save(file_name)

    return (status,errtext,file_name)
# ed def